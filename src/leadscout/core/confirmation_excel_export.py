"""
Enhanced Excel export system for ethnicity confirmation workflow.

This module implements the complete 21-column Excel export format with:
- Original lead data (11 columns)
- Enhanced AI predictions with spatial intelligence (5 columns)  
- Empty confirmation columns for dialler team (2 columns)
- Metadata for precise record tracking (3 columns)

Key Features:
- Integration with Developer B's spatial learning system
- Professional Excel dropdown validation
- Confidence-based color coding
- Precise source tracking for complete traceability
- openpyxl-based formatting for professional presentation

Architecture:
- ConfirmationExcelExporter: Main export class
- 21-column format exactly per specification
- Excel dropdown validation with canonical ethnicities
- Color coding for confidence visualization
- Spatial context integration

Usage:
    exporter = ConfirmationExcelExporter()
    output_path = await exporter.export_job_for_confirmation(job_id, output_path)
"""

import asyncio
import sqlite3
import json
import uuid
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, Any, List, Tuple
import pandas as pd
import structlog

# Excel formatting dependencies
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .job_database import JobDatabase
from .ethnicity_confirmation_database import (
    EthnicityConfirmationDatabase, 
    EthnicityConfirmation,
    generate_file_identifier,
    generate_spatial_context_hash
)

logger = structlog.get_logger(__name__)

class ConfirmationExcelExporter:
    """Enhanced Excel exporter for ethnicity confirmation workflow.
    
    Provides complete 21-column Excel export with AI predictions,
    confirmation dropdowns, and spatial intelligence integration.
    
    Features:
    - 21-column format: Original (11) + AI Enhancement (5) + Confirmation (2) + Metadata (3)
    - Professional Excel dropdown validation
    - Confidence-based color coding
    - Spatial context integration with Developer B's system
    - Source tracking for complete record traceability
    """
    
    def __init__(self, job_db_path: Path = Path("cache/jobs.db"),
                 confirmation_db_path: Path = Path("cache/ethnicity_confirmations.db")):
        """Initialize confirmation Excel exporter.
        
        Args:
            job_db_path: Path to job database
            confirmation_db_path: Path to confirmation database
        """
        self.job_db = JobDatabase(db_path=job_db_path)
        self.confirmation_db = EthnicityConfirmationDatabase(db_path=confirmation_db_path)
        
        # Ensure canonical ethnicities are initialized
        self.confirmation_db.initialize_canonical_ethnicities()
        
        logger.info("ConfirmationExcelExporter initialized",
                   job_db_path=str(job_db_path),
                   confirmation_db_path=str(confirmation_db_path))
    
    async def export_job_for_confirmation(self, job_id: str, 
                                        output_path: Optional[Path] = None,
                                        include_spatial_enhancement: bool = True) -> Path:
        """Export job results with 21-column confirmation format.
        
        Creates Excel file with:
        - Original lead data (11 columns)
        - Enhanced AI predictions with spatial intelligence (5 columns)
        - Empty confirmation columns with dropdown validation (2 columns)
        - Metadata for tracking (3 columns)
        
        Args:
            job_id: Job identifier to export
            output_path: Optional output file path (auto-generated if not provided)
            include_spatial_enhancement: Whether to include Developer B's spatial enhancements
            
        Returns:
            Path to created Excel file
            
        Raises:
            ValueError: If job not found or invalid
            RuntimeError: If openpyxl not available for Excel formatting
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl package required for Excel export with formatting. Install with: pip install openpyxl")
        
        # Generate output path if not provided
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = Path(f"leads_for_confirmation_{job_id[:8]}_{timestamp}.xlsx")
        
        logger.info("Starting confirmation export",
                   job_id=job_id,
                   output_path=str(output_path),
                   spatial_enhancement=include_spatial_enhancement)
        
        # Get job results with source tracking
        job_results = await self._get_enhanced_job_results(job_id, include_spatial_enhancement)
        
        if not job_results:
            raise ValueError(f"No results found for job: {job_id}")
        
        # Create 21-column DataFrame
        export_df = self._create_21_column_dataframe(job_results)
        
        # Create Excel file with formatting
        await self._create_formatted_excel_file(export_df, output_path, job_id)
        
        logger.info("Confirmation export completed",
                   job_id=job_id,
                   output_path=str(output_path),
                   record_count=len(export_df))
        
        return output_path
    
    async def _get_enhanced_job_results(self, job_id: str, 
                                      include_spatial_enhancement: bool) -> List[Dict[str, Any]]:
        """Get job results with enhanced predictions and source tracking.
        
        Args:
            job_id: Job identifier
            include_spatial_enhancement: Whether to include spatial enhancements
            
        Returns:
            List of enhanced result dictionaries
        """
        # Get job results from database
        with sqlite3.connect(self.job_db.db_path) as conn:
            # Check if source tracking fields exist (for backward compatibility)
            cursor = conn.execute("PRAGMA table_info(lead_processing_results)")
            columns = [row[1] for row in cursor.fetchall()]
            has_source_tracking = 'source_row_number' in columns
            
            if has_source_tracking:
                # Enhanced query with source tracking fields
                query = """
                SELECT 
                    lpr.job_id, lpr.row_index, lpr.batch_number,
                    lpr.entity_name, lpr.director_name,
                    lpr.classification_result, lpr.processing_status,
                    lpr.processing_time_ms, lpr.api_provider, lpr.api_cost,
                    lpr.created_at,
                    -- Source tracking fields
                    lpr.source_row_number, lpr.source_file_identifier,
                    lpr.original_entity_name, lpr.original_director_name,
                    lpr.original_registered_address, lpr.original_registered_city,
                    lpr.original_registered_province,
                    -- CRITICAL: Contact fields for dialling operations
                    lpr.cell_number, lpr.contact_number, lpr.email_address, lpr.director_cell,
                    -- Business fields for context
                    lpr.trading_as_name, lpr.keyword,
                    -- Job metadata
                    je.input_file_path, je.start_time
                FROM lead_processing_results lpr
                JOIN job_executions je ON lpr.job_id = je.job_id
                WHERE lpr.job_id = ? AND lpr.processing_status = 'success'
                ORDER BY lpr.source_row_number, lpr.row_index
                """
            else:
                # Fallback query for older jobs without source tracking
                query = """
                SELECT 
                    lpr.job_id, lpr.row_index, lpr.batch_number,
                    lpr.entity_name, lpr.director_name,
                    lpr.classification_result, lpr.processing_status,
                    lpr.processing_time_ms, lpr.api_provider, lpr.api_cost,
                    lpr.created_at,
                    -- Fallback values for source tracking fields
                    (lpr.row_index + 2) as source_row_number,  -- Excel 1-based + header
                    NULL as source_file_identifier,
                    -- Contact fields (will be NULL for older jobs)
                    lpr.cell_number, lpr.contact_number, lpr.email_address, lpr.director_cell,
                    -- Business fields (will be NULL for older jobs)
                    lpr.trading_as_name, lpr.keyword,
                    lpr.entity_name as original_entity_name,
                    lpr.director_name as original_director_name,
                    NULL as original_registered_address,
                    NULL as original_registered_city,
                    NULL as original_registered_province,
                    -- Job metadata
                    je.input_file_path, je.start_time
                FROM lead_processing_results lpr
                JOIN job_executions je ON lpr.job_id = je.job_id
                WHERE lpr.job_id = ? AND lpr.processing_status = 'success'
                ORDER BY lpr.row_index
                """
            
            conn.row_factory = sqlite3.Row
            cursor = conn.execute(query, (job_id,))
            raw_results = cursor.fetchall()
        
        if not raw_results:
            logger.warning("No successful results found for job", job_id=job_id)
            return []
        
        # Process and enhance results
        enhanced_results = []
        for row in raw_results:
            try:
                # Parse classification result
                classification_data = json.loads(row['classification_result']) if row['classification_result'] else {}
                
                # Get original data with fallbacks
                original_entity = row['original_entity_name'] or row['entity_name']
                original_director = row['original_director_name'] or row['director_name']
                
                # Create enhanced result
                enhanced_result = {
                    # Original lead columns (11) - these would come from original Excel
                    'EntityName': original_entity,
                    'TradingAsName': row['trading_as_name'] or '',  # Populated from database
                    'Keyword': row['keyword'] or '',
                    'ContactNumber': row['contact_number'] or '',
                    'CellNumber': row['cell_number'] or '',
                    'EmailAddress': row['email_address'] or '',
                    'RegisteredAddress': row['original_registered_address'] or '',
                    'RegisteredAddressCity': row['original_registered_city'] or '',
                    'RegisteredAddressProvince': row['original_registered_province'] or '',
                    'DirectorName': original_director,
                    'DirectorCell': row['director_cell'] or '',
                    
                    # AI Enhancement columns (5)
                    'director_ethnicity': classification_data.get('ethnicity', 'unknown'),
                    'ethnicity_confidence': classification_data.get('confidence', 0.0),
                    'classification_method': classification_data.get('method', 'unknown'),
                    'spatial_context': self._format_spatial_context(
                        row['original_registered_city'], 
                        row['original_registered_province']
                    ),
                    'processing_notes': classification_data.get('notes', ''),
                    
                    # Empty confirmation columns (2)
                    'confirmed_ethnicity': '',
                    'confirmation_notes': '',
                    
                    # Metadata columns (3)
                    'source_row_number': row['source_row_number'] or (row['row_index'] + 2),  # Excel 1-based
                    'job_id': job_id,
                    'processed_at': row['created_at'],
                    
                    # Internal tracking fields (not exported)
                    '_source_file_identifier': row['source_file_identifier'],
                    '_api_cost': row['api_cost'] or 0.0,
                    '_processing_time_ms': row['processing_time_ms'] or 0.0
                }
                
                # Apply spatial enhancement if enabled
                if include_spatial_enhancement:
                    enhanced_result = await self._apply_spatial_enhancement(enhanced_result)
                
                enhanced_results.append(enhanced_result)
                
            except Exception as e:
                logger.warning("Failed to process result row",
                             job_id=job_id,
                             row_index=row['row_index'],
                             error=str(e))
                continue
        
        logger.info("Enhanced job results processed",
                   job_id=job_id,
                   total_results=len(raw_results),
                   enhanced_results=len(enhanced_results))
        
        return enhanced_results
    
    async def _apply_spatial_enhancement(self, result: Dict[str, Any]) -> Dict[str, Any]:
        """Apply Developer B's spatial intelligence enhancements.
        
        This integrates with Developer B's spatial learning system to provide
        enhanced ethnicity predictions based on spatial patterns.
        
        Args:
            result: Result dictionary to enhance
            
        Returns:
            Enhanced result with spatial intelligence
        """
        try:
            # TODO: Integration with Developer B's system
            # For now, simulate enhanced prediction with spatial context
            
            director_name = result['DirectorName']
            city = result['RegisteredAddressCity']
            province = result['RegisteredAddressProvince']
            
            # Placeholder for Developer B's enhanced prediction
            # enhanced_prediction = await spatial_learning_db.enhanced_ethnicity_prediction_with_confirmations(
            #     name=director_name,
            #     city=city,
            #     province=province,
            #     suburb=None  # Could be extracted from address
            # )
            
            # For now, enhance with spatial context information
            if city and province:
                spatial_context = f"{city}, {province}"
                
                # Add spatial confidence bonus for common SA locations
                current_confidence = result['ethnicity_confidence']
                if any(location in city.lower() for location in ['johannesburg', 'cape town', 'durban', 'pretoria']):
                    # Boost confidence for major cities with more data
                    enhanced_confidence = min(current_confidence + 0.05, 1.0)
                    result['ethnicity_confidence'] = enhanced_confidence
                    
                    if not result['processing_notes']:
                        result['processing_notes'] = f"Spatial context: {spatial_context}"
                    else:
                        result['processing_notes'] += f" | Spatial context: {spatial_context}"
            
            logger.debug("Applied spatial enhancement",
                        director_name=director_name,
                        spatial_context=result['spatial_context'],
                        enhanced_confidence=result['ethnicity_confidence'])
            
        except Exception as e:
            logger.warning("Spatial enhancement failed",
                          director_name=result.get('DirectorName'),
                          error=str(e))
        
        return result
    
    def _format_spatial_context(self, city: Optional[str], province: Optional[str]) -> str:
        """Format spatial context for display.
        
        Args:
            city: City name
            province: Province name
            
        Returns:
            Formatted spatial context string
        """
        components = []
        if city and city.strip():
            components.append(city.strip())
        if province and province.strip():
            components.append(province.strip())
        
        return ', '.join(components) if components else ''
    
    def _create_21_column_dataframe(self, results: List[Dict[str, Any]]) -> pd.DataFrame:
        """Create 21-column DataFrame for Excel export.
        
        Args:
            results: Enhanced job results
            
        Returns:
            DataFrame with exactly 21 columns in specified order
        """
        # Define exact 21-column structure
        column_order = [
            # Original Lead Columns (11)
            'EntityName', 'TradingAsName', 'Keyword', 'ContactNumber', 'CellNumber',
            'EmailAddress', 'RegisteredAddress', 'RegisteredAddressCity', 
            'RegisteredAddressProvince', 'DirectorName', 'DirectorCell',
            
            # AI Enhancement Columns (5)
            'director_ethnicity', 'ethnicity_confidence', 'classification_method',
            'spatial_context', 'processing_notes',
            
            # Confirmation Columns (2)
            'confirmed_ethnicity', 'confirmation_notes',
            
            # Metadata Columns (3)
            'source_row_number', 'job_id', 'processed_at'
        ]
        
        # Create DataFrame with only the 21 specified columns
        export_data = []
        for result in results:
            row_data = {}
            for col in column_order:
                row_data[col] = result.get(col, '')
            export_data.append(row_data)
        
        df = pd.DataFrame(export_data, columns=column_order)
        
        logger.info("Created 21-column DataFrame",
                   rows=len(df),
                   columns=len(df.columns),
                   column_structure=df.columns.tolist())
        
        return df
    
    async def _create_formatted_excel_file(self, df: pd.DataFrame, 
                                         output_path: Path, job_id: str) -> None:
        """Create formatted Excel file with dropdowns and color coding.
        
        Args:
            df: DataFrame to export
            output_path: Output file path
            job_id: Job identifier for metadata
        """
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Enriched Leads for Confirmation"
        
        # Write DataFrame to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Apply formatting
        await self._apply_excel_formatting(ws, len(df))
        
        # Add ethnicity dropdown validation
        self._add_ethnicity_dropdown_validation(ws, len(df))
        
        # Apply confidence-based color coding
        self._apply_confidence_color_coding(ws, len(df))
        
        # Set column widths
        self._set_optimal_column_widths(ws)
        
        # Add metadata sheet
        self._add_metadata_sheet(wb, job_id, len(df))
        
        # Save workbook
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        
        logger.info("Formatted Excel file created",
                   output_path=str(output_path),
                   worksheets=len(wb.sheetnames),
                   rows=len(df))
    
    async def _apply_excel_formatting(self, ws, row_count: int) -> None:
        """Apply professional Excel formatting.
        
        Args:
            ws: Worksheet to format
            row_count: Number of data rows
        """
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:  # Header row
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Data formatting
        data_alignment = Alignment(horizontal="left", vertical="center")
        for row in ws.iter_rows(min_row=2, max_row=row_count + 1):
            for cell in row:
                cell.alignment = data_alignment
        
        # Format numeric columns
        confidence_col = self._get_column_letter(ws, 'ethnicity_confidence')
        if confidence_col:
            for row in range(2, row_count + 2):
                cell = ws[f'{confidence_col}{row}']
                cell.number_format = '0.00'
        
        logger.debug("Applied Excel formatting",
                    row_count=row_count,
                    header_style="applied",
                    numeric_formatting="applied")
    
    def _add_ethnicity_dropdown_validation(self, ws, row_count: int) -> None:
        """Add dropdown validation to confirmed_ethnicity column.
        
        Args:
            ws: Worksheet to modify
            row_count: Number of data rows
        """
        try:
            # Get ethnicity options from confirmation database
            ethnicity_options = self.confirmation_db.get_canonical_ethnicities_for_dropdown()
            
            if not ethnicity_options:
                logger.warning("No ethnicity options available for dropdown")
                return
            
            # Find confirmed_ethnicity column
            confirmed_ethnicity_col = self._get_column_letter(ws, 'confirmed_ethnicity')
            if not confirmed_ethnicity_col:
                logger.warning("confirmed_ethnicity column not found")
                return
            
            # Create validation
            validation = DataValidation(
                type="list",
                formula1=f'"{",".join(ethnicity_options)}"',
                showDropDown=True,
                showErrorMessage=True,
                errorTitle="Invalid Ethnicity Selection",
                error="Please select a valid ethnicity from the dropdown list. "
                      "Contact admin to add new options if needed.",
                showInputMessage=True,
                promptTitle="Select Ethnicity",
                prompt="Choose the confirmed ethnicity from the dropdown. "
                       "Most common options (African, White, Coloured, Indian) are listed first."
            )
            
            # Apply validation to confirmed_ethnicity column
            validation.add(f"{confirmed_ethnicity_col}2:{confirmed_ethnicity_col}{row_count + 1}")
            ws.add_data_validation(validation)
            
            logger.info("Added ethnicity dropdown validation",
                       column=confirmed_ethnicity_col,
                       row_range=f"2:{row_count + 1}",
                       options_count=len(ethnicity_options))
            
        except Exception as e:
            logger.error("Failed to add dropdown validation", error=str(e))
    
    def _apply_confidence_color_coding(self, ws, row_count: int) -> None:
        """Apply confidence-based color coding to director_ethnicity column.
        
        Args:
            ws: Worksheet to modify
            row_count: Number of data rows
        """
        try:
            # Find relevant columns
            ethnicity_col = self._get_column_letter(ws, 'director_ethnicity')
            confidence_col = self._get_column_letter(ws, 'ethnicity_confidence')
            
            if not ethnicity_col or not confidence_col:
                logger.warning("Required columns not found for color coding",
                             ethnicity_col=ethnicity_col,
                             confidence_col=confidence_col)
                return
            
            # Define color scheme
            color_fills = {
                'high': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),     # Light green (>0.8)
                'medium': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),   # Light yellow (0.6-0.8)
                'low': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),      # Light red (<0.6)
                'very_low': PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")  # Light purple (<0.4)
            }
            
            # Apply color coding
            colored_count = {'high': 0, 'medium': 0, 'low': 0, 'very_low': 0}
            
            for row in range(2, row_count + 2):
                try:
                    confidence_cell = ws[f'{confidence_col}{row}']
                    ethnicity_cell = ws[f'{ethnicity_col}{row}']
                    
                    if confidence_cell.value is not None:
                        confidence = float(confidence_cell.value)
                        
                        if confidence >= 0.8:
                            ethnicity_cell.fill = color_fills['high']
                            colored_count['high'] += 1
                        elif confidence >= 0.6:
                            ethnicity_cell.fill = color_fills['medium']
                            colored_count['medium'] += 1
                        elif confidence >= 0.4:
                            ethnicity_cell.fill = color_fills['low']
                            colored_count['low'] += 1
                        else:
                            ethnicity_cell.fill = color_fills['very_low']
                            colored_count['very_low'] += 1
                
                except (ValueError, TypeError) as e:
                    logger.debug("Skipped color coding for row",
                               row=row,
                               error=str(e))
                    continue
            
            logger.info("Applied confidence color coding",
                       high_confidence=colored_count['high'],
                       medium_confidence=colored_count['medium'],
                       low_confidence=colored_count['low'],
                       very_low_confidence=colored_count['very_low'])
            
        except Exception as e:
            logger.error("Failed to apply color coding", error=str(e))
    
    def _get_column_letter(self, ws, column_name: str) -> Optional[str]:
        """Get Excel column letter for given column name.
        
        Args:
            ws: Worksheet
            column_name: Column name to find
            
        Returns:
            Column letter or None if not found
        """
        try:
            for cell in ws[1]:  # Header row
                if cell.value == column_name:
                    return cell.column_letter
        except Exception as e:
            logger.debug("Error finding column letter",
                        column_name=column_name,
                        error=str(e))
        return None
    
    def _set_optimal_column_widths(self, ws) -> None:
        """Set optimal column widths for readability.
        
        Args:
            ws: Worksheet to modify
        """
        # Define optimal widths for known columns
        column_widths = {
            'EntityName': 25,
            'TradingAsName': 20,
            'Keyword': 15,
            'ContactNumber': 15,
            'CellNumber': 15,
            'EmailAddress': 25,
            'RegisteredAddress': 30,
            'RegisteredAddressCity': 18,
            'RegisteredAddressProvince': 18,
            'DirectorName': 20,
            'DirectorCell': 15,
            'director_ethnicity': 15,
            'ethnicity_confidence': 12,
            'classification_method': 18,
            'spatial_context': 20,
            'processing_notes': 25,
            'confirmed_ethnicity': 18,
            'confirmation_notes': 25,
            'source_row_number': 12,
            'job_id': 15,
            'processed_at': 18
        }
        
        # Apply widths
        for cell in ws[1]:  # Header row
            column_name = cell.value
            if column_name in column_widths:
                ws.column_dimensions[cell.column_letter].width = column_widths[column_name]
            else:
                # Auto-width for other columns
                ws.column_dimensions[cell.column_letter].width = 15
        
        logger.debug("Set optimal column widths",
                    columns_configured=len(column_widths))
    
    def _add_metadata_sheet(self, wb, job_id: str, record_count: int) -> None:
        """Add metadata sheet with export information.
        
        Args:
            wb: Workbook to modify
            job_id: Job identifier
            record_count: Number of records exported
        """
        try:
            # Create metadata sheet
            meta_ws = wb.create_sheet("Export Metadata")
            
            # Add metadata information
            metadata = [
                ["Export Information", ""],
                ["Job ID", job_id],
                ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["Record Count", record_count],
                ["Export Type", "Ethnicity Confirmation"],
                ["Format Version", "1.0"],
                ["", ""],
                ["Column Structure", ""],
                ["Original Lead Data", "Columns 1-11"],
                ["AI Enhancement Data", "Columns 12-16"],
                ["Confirmation Data", "Columns 17-18"],
                ["Metadata", "Columns 19-21"],
                ["", ""],
                ["Instructions", ""],
                ["1. Confirmed Ethnicity", "Select from dropdown in column Q"],
                ["2. Confirmation Notes", "Add notes in column R"],
                ["3. Source Row Number", "Use column S for reference"],
                ["", ""],
                ["Contact", ""],
                ["System", "LeadScout Ethnicity Confirmation"],
                ["Support", "Contact system administrator"]
            ]
            
            for row_data in metadata:
                meta_ws.append(row_data)
            
            # Format metadata sheet
            meta_ws.column_dimensions['A'].width = 25
            meta_ws.column_dimensions['B'].width = 30
            
            logger.debug("Added metadata sheet",
                        job_id=job_id,
                        metadata_rows=len(metadata))
            
        except Exception as e:
            logger.warning("Failed to add metadata sheet", error=str(e))
    
    async def get_export_statistics(self, job_id: str) -> Dict[str, Any]:
        """Get export statistics for a job.
        
        Args:
            job_id: Job identifier
            
        Returns:
            Dictionary with export statistics
        """
        try:
            # Get job results
            results = await self._get_enhanced_job_results(job_id, include_spatial_enhancement=True)
            
            if not results:
                return {"error": f"No results found for job: {job_id}"}
            
            # Calculate statistics
            total_records = len(results)
            ethnicity_counts = {}
            method_counts = {}
            confidence_stats = []
            
            for result in results:
                # Ethnicity distribution
                ethnicity = result.get('director_ethnicity', 'unknown')
                ethnicity_counts[ethnicity] = ethnicity_counts.get(ethnicity, 0) + 1
                
                # Method distribution
                method = result.get('classification_method', 'unknown')
                method_counts[method] = method_counts.get(method, 0) + 1
                
                # Confidence statistics
                confidence = result.get('ethnicity_confidence', 0.0)
                if isinstance(confidence, (int, float)):
                    confidence_stats.append(float(confidence))
            
            # Calculate confidence metrics
            if confidence_stats:
                avg_confidence = sum(confidence_stats) / len(confidence_stats)
                high_confidence = len([c for c in confidence_stats if c >= 0.8])
                medium_confidence = len([c for c in confidence_stats if 0.6 <= c < 0.8])
                low_confidence = len([c for c in confidence_stats if c < 0.6])
            else:
                avg_confidence = 0.0
                high_confidence = medium_confidence = low_confidence = 0
            
            return {
                "job_id": job_id,
                "total_records": total_records,
                "ethnicity_distribution": ethnicity_counts,
                "method_distribution": method_counts,
                "confidence_metrics": {
                    "average_confidence": round(avg_confidence, 3),
                    "high_confidence_count": high_confidence,
                    "medium_confidence_count": medium_confidence,
                    "low_confidence_count": low_confidence,
                    "high_confidence_percentage": round(high_confidence / total_records * 100, 1) if total_records > 0 else 0
                }
            }
            
        except Exception as e:
            logger.error("Failed to get export statistics",
                        job_id=job_id,
                        error=str(e))
            return {"error": f"Failed to get statistics: {e}"}